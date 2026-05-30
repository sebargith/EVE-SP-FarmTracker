import csv
import os
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from src.calculations.scenarios import generate_scenario_matrix


WORKBOOK_PATH = Path(__file__).resolve().parents[1] / "eve_sp_farm_plex_only_profit_model.xlsx"
FIXTURE_PATH = Path(__file__).resolve().parent / "fixtures" / "reference_scenarios.csv"

LABEL_COLUMNS = [
    "Scenario Type",
    "Training / Omega Offer",
    "MCT Source",
    "Extractor Source",
    "Status",
]

NUMERIC_COLUMNS = [
    "Omega PLEX Total",
    "Omega Months",
    "Queues",
    "MCT Unit PLEX",
    "MCT Market Unit ISK",
    "MCT PLEX Total",
    "MCT Market ISK Total",
    "Bundle PLEX Total",
    "Training PLEX Total",
    "Training Cost ISK",
    "Queue-Months",
    "Training SP",
    "Bonus SP",
    "Total SP",
    "Injectors Produced",
    "Full Injectors",
    "Remainder SP",
    "Extractor Unit PLEX",
    "Extractor Unit ISK",
    "Extractor Cost ISK",
    "LSI Gross Revenue",
    "Market Fees",
    "LSI Net Revenue",
    "Total Cost ISK",
    "Profit ISK",
    "Profit / Calendar Month ISK",
    "Profit / Queue-Month ISK",
    "Break-even LSI Sell Price",
]


def test_generated_scenarios_match_locked_reference_fixture() -> None:
    fixture_rows = _load_fixture_rows()
    generated = generate_scenario_matrix().set_index("ID")

    assert len(fixture_rows) == 11

    for scenario_id, expected in fixture_rows.items():
        actual = generated.loc[scenario_id]
        _assert_row_matches(scenario_id, actual, expected)


def test_generated_scenarios_match_reference_workbook_core_outputs_audit() -> None:
    if os.getenv("RUN_WORKBOOK_AUDIT") != "1":
        pytest.skip("Set RUN_WORKBOOK_AUDIT=1 to compare against Excel cached values.")

    workbook_rows = _load_workbook_rows()
    generated = generate_scenario_matrix().set_index("ID")

    assert len(workbook_rows) == 927
    assert len(generated) == 927

    for scenario_id, expected in workbook_rows.items():
        actual = generated.loc[scenario_id]
        _assert_row_matches(scenario_id, actual, expected)


def _assert_row_matches(
    scenario_id: int,
    actual: Any,
    expected: dict[str, Any],
) -> None:
    for column in LABEL_COLUMNS:
        assert actual[column] == expected[column], f"{scenario_id}: {column}"

    for column in NUMERIC_COLUMNS:
        assert float(actual[column]) == pytest.approx(
            float(expected[column]),
            rel=1e-10,
            abs=0.01,
        ), f"{scenario_id}: {column}"


def _load_fixture_rows() -> dict[int, dict[str, Any]]:
    with FIXTURE_PATH.open("r", encoding="utf-8", newline="") as file:
        reader = csv.DictReader(file)
        return {int(row["ID"]): dict(row) for row in reader}


def _load_workbook_rows() -> dict[int, dict[str, Any]]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=False)
    worksheet = workbook["02 Scenario Matrix"]
    headers = [worksheet.cell(2, column).value for column in range(1, worksheet.max_column + 1)]
    column_indexes = {header: index + 1 for index, header in enumerate(headers)}

    rows: dict[int, dict[str, Any]] = {}
    for row_index in range(3, worksheet.max_row + 1):
        scenario_id = worksheet.cell(row_index, column_indexes["ID"]).value
        rows[int(scenario_id)] = {
            column: worksheet.cell(row_index, column_indexes[column]).value
            for column in [*LABEL_COLUMNS, *NUMERIC_COLUMNS]
        }

    return rows
